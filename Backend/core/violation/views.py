from datetime import datetime, timedelta
from calendar import monthrange
from django.db.models import Count
from django.db.models.functions import TruncDay
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import viewsets
from rest_framework.generics import get_object_or_404
from rest_framework.pagination import PageNumberPagination

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Violation
from .serializers import ViolationSerializer


# ── Helpers ────────────────────────────────────────────────────────────────────
def _parse_month_year(month_year: str):
    """
    Parse 'March_2026' → (start_date, end_date, display_str)
    Raises ValueError on bad input.
    """
    month_date = datetime.strptime(month_year, "%B_%Y")
    start = month_date.replace(day=1).date()
    last  = monthrange(month_date.year, month_date.month)[1]
    end   = month_date.replace(day=last).date()
    return start, end, start.strftime("%B %Y")


def _daily_series(qs, start, end):
    """
    Return list of { day: int, total: int } for every day in [start, end].
    qs must already be filtered to the right month.
    """
    raw = (
        qs.annotate(trunc_day=TruncDay("time"))
          .values("trunc_day")
          .annotate(total=Count("id"))
          .order_by("trunc_day")
    )
    data_map = {item["trunc_day"].day: item["total"] for item in raw}

    today      = timezone.now().date()
    is_current = (start.year == today.year and start.month == today.month)
    last_day   = today.day if is_current else end.day

    return [
        {"day": d, "total": data_map.get(d, 0)}
        for d in range(1, last_day + 1)
    ]


# ── Custom Paginator ───────────────────────────────────────────────────────────
class ViolationPagination(PageNumberPagination):
    page_size             = 5
    page_size_query_param = "page_size"
    max_page_size         = 100


# ── Base CRUD ──────────────────────────────────────────────────────────────────
class ViolationViewSet(viewsets.ModelViewSet):
    queryset         = Violation.objects.all().order_by("-time")
    serializer_class = ViolationSerializer
    pagination_class = ViolationPagination


# ══════════════════════════════════════════════════════════════════════════════
# Analytics — summary (existing, now supports ?month= filter)
# GET /violations/analytics/?month=March_2026
# ══════════════════════════════════════════════════════════════════════════════
class ViolationAnalyticsView(APIView):

    def get(self, request):
        month_param = request.query_params.get("month")
        qs = Violation.objects.all()

        if month_param:
            try:
                start, end, _ = _parse_month_year(month_param)
                qs = qs.filter(time__date__gte=start, time__date__lte=end)
            except ValueError:
                return Response(
                    {"error": "Invalid month format. Use Month_Year e.g. March_2026"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        total_violations = qs.count()

        model_wise = (
            qs.values("ml_model__id", "ml_model__name")
              .annotate(total=Count("id"))
              .order_by("-total")
        )

        all_detections = (
            qs.exclude(detections=[])
              .values_list("detections", flat=True)
        )

        label_counts = {}
        for detection_list in all_detections:
            for det in detection_list:
                label = det.get("label", "unknown")
                label_counts[label] = label_counts.get(label, 0) + 1

        top_labels = sorted(
            [{"label": k, "count": v} for k, v in label_counts.items()],
            key=lambda x: x["count"],
            reverse=True,
        )

        top_vehicles = (
            qs.exclude(plate_number__isnull=True)
              .exclude(plate_number="")
              .values("plate_number")
              .annotate(total=Count("id"))
              .order_by("-total")[:10]
        )

        return Response({
            "total_violations":         total_violations,
            "ml_model_wise_violations": list(model_wise),
            "top_detected_labels":      top_labels,
            "top_vehicles":             list(top_vehicles),
        })


# ══════════════════════════════════════════════════════════════════════════════
# ✅ NEW — Per-model analytics (used by Model Comparison panel)
# GET /violations/analytics/model/<model_id>/?month=March_2026
# Response: { total_violations, daily_trends: [{day, total}] }
# ══════════════════════════════════════════════════════════════════════════════
class ViolationModelAnalyticsView(APIView):

    def get(self, request, model_id):
        month_param = request.query_params.get("month")

        qs = Violation.objects.filter(ml_model__id=model_id)

        if month_param:
            try:
                start, end, display = _parse_month_year(month_param)
            except ValueError:
                return Response(
                    {"error": "Invalid month format. Use Month_Year e.g. March_2026"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            today = timezone.now().date()
            start = today.replace(day=1)
            last  = monthrange(today.year, today.month)[1]
            end   = today.replace(day=last)
            display = today.strftime("%B %Y")

        qs = qs.filter(time__date__gte=start, time__date__lte=end)

        return Response({
            "model_id":        model_id,
            "month":           display,
            "total_violations": qs.count(),
            "daily_trends":    _daily_series(qs, start, end),
        })


# ══════════════════════════════════════════════════════════════════════════════
# ✅ NEW — Camera-wise comparison (used by Camera Comparison panel)
# GET /violations/analytics/camera-wise/?cameras=1,2,3&month=March_2026
# Response: { cameras: [{ camera_id, camera_name, total_violations,
#                         daily_trends: [{day, total}] }] }
# ══════════════════════════════════════════════════════════════════════════════
class ViolationCameraWiseView(APIView):

    def get(self, request):
        cameras_param = request.query_params.get("cameras", "")
        month_param   = request.query_params.get("month")

        # Parse camera IDs
        try:
            camera_ids = [
                int(cid.strip())
                for cid in cameras_param.split(",")
                if cid.strip().isdigit()
            ]
        except ValueError:
            camera_ids = []

        if not camera_ids:
            return Response(
                {"error": "Provide at least one camera id via ?cameras=1,2,3"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parse month
        if month_param:
            try:
                start, end, display = _parse_month_year(month_param)
            except ValueError:
                return Response(
                    {"error": "Invalid month format. Use Month_Year e.g. March_2026"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            today   = timezone.now().date()
            start   = today.replace(day=1)
            last    = monthrange(today.year, today.month)[1]
            end     = today.replace(day=last)
            display = today.strftime("%B %Y")

        result = []
        for cam_id in camera_ids:
            qs = Violation.objects.filter(
                camera__id=cam_id,
                time__date__gte=start,
                time__date__lte=end,
            )

            # Resolve a human-readable camera name
            first = qs.select_related("camera").first()
            camera_name = str(first.camera) if first else f"Camera {cam_id}"

            result.append({
                "camera_id":        cam_id,
                "camera_name":      camera_name,
                "total_violations": qs.count(),
                "daily_trends":     _daily_series(qs, start, end),
            })

        return Response({
            "month":   display,
            "cameras": result,
        })


# ══════════════════════════════════════════════════════════════════════════════
# ✅ NEW — Location-wise analytics (used by Location Comparison panel)
# GET /violations/analytics/location-wise/?month=March_2026&top=10
# Response: { locations: [{ location, total_violations }] }
# ══════════════════════════════════════════════════════════════════════════════
class ViolationLocationWiseView(APIView):

    def get(self, request):
        month_param = request.query_params.get("month")
        top_n       = int(request.query_params.get("top", 10))

        qs = Violation.objects.all()

        if month_param:
            try:
                start, end, display = _parse_month_year(month_param)
                qs = qs.filter(time__date__gte=start, time__date__lte=end)
            except ValueError:
                return Response(
                    {"error": "Invalid month format. Use Month_Year e.g. March_2026"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            today   = timezone.now().date()
            start   = today.replace(day=1)
            last    = monthrange(today.year, today.month)[1]
            end     = today.replace(day=last)
            display = today.strftime("%B %Y")
            qs = qs.filter(time__date__gte=start, time__date__lte=end)

        location_wise = (
            qs.values("camera__location")          # uses Camera.location field
              .annotate(total_violations=Count("id"))
              .order_by("-total_violations")[:top_n]
        )

        locations = [
            {
                "location":         item["camera__location"] or "Unknown",
                "total_violations": item["total_violations"],
            }
            for item in location_wise
        ]

        return Response({
            "month":     display,
            "top":       top_n,
            "locations": locations,
        })


# ── Monthly Trends ─────────────────────────────────────────────────────────────
class ViolationMonthlyTrendsView(APIView):
    """GET /violations/monthly-trends/<month_year>/"""

    def get(self, request, month_year):
        try:
            start, end, display = _parse_month_year(month_year)
        except ValueError:
            return Response(
                {"error": "Invalid format. Use Month_Year e.g. February_2026"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = Violation.objects.filter(
            time__date__gte=start,
            time__date__lte=end,
        )

        trends = [
            {"date": start.replace(day=d["day"]).strftime("%Y-%m-%d"), **d}
            for d in _daily_series(qs, start, end)
        ]

        return Response({
            "monthly_trends": trends,
            "current_month":  display,
        })


# ── Per-violation Detections ───────────────────────────────────────────────────
class ViolationDetectionsView(APIView):
    """GET /violations/<pk>/detections/"""

    def get(self, request, pk=None):
        violation = get_object_or_404(Violation, pk=pk)
        return Response({
            "violation_id":  violation.id,
            "time":          violation.time,
            "plate_number":  violation.plate_number,
            "detections":    violation.detections,
            "total_objects": len(violation.detections),
        })


# ── Violation Count Stats ──────────────────────────────────────────────────────
class ViolationCountStatsView(APIView):
    """GET /violations/count-stats/"""

    def get(self, request):
        now   = timezone.now()
        today = now.date()

        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end   = timezone.make_aware(datetime.combine(today, datetime.max.time()))

        seven_days_ago  = today - timedelta(days=7)
        yesterday       = today - timedelta(days=1)
        last_week_start = timezone.make_aware(datetime.combine(seven_days_ago, datetime.min.time()))
        last_week_end   = timezone.make_aware(datetime.combine(yesterday,      datetime.max.time()))

        month_start    = today.replace(day=1)
        month_start_dt = timezone.make_aware(datetime.combine(month_start, datetime.min.time()))

        return Response({
            "today": {
                "count":      Violation.objects.filter(time__gte=today_start, time__lte=today_end).count(),
                "date":       today.strftime("%Y-%m-%d"),
                "date_range": {"start": today_start.isoformat(), "end": today_end.isoformat()},
            },
            "last_week": {
                "count":       Violation.objects.filter(time__gte=last_week_start, time__lte=last_week_end).count(),
                "days":        7,
                "date_range":  {"start": last_week_start.isoformat(), "end": last_week_end.isoformat()},
                "description": "Last 7 days excluding today",
            },
            "current_month": {
                "count":      Violation.objects.filter(time__gte=month_start_dt, time__lte=today_end).count(),
                "month":      today.strftime("%B %Y"),
                "date_range": {"start": month_start_dt.isoformat(), "end": today_end.isoformat()},
            },
            "timestamp": now.isoformat(),
        })


# ── Excel Export ───────────────────────────────────────────────────────────────
class ViolationExcelExportView(APIView):
    """GET /violations/export/excel/?period=today|last_week|current_month|all"""

    def get(self, request):
        period         = request.query_params.get("period", "all")
        start_date_str = request.query_params.get("start_date")
        end_date_str   = request.query_params.get("end_date")

        now      = timezone.now()
        today    = now.date()
        queryset = Violation.objects.all()

        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date   = datetime.strptime(end_date_str,   "%Y-%m-%d").date()
                queryset   = queryset.filter(
                    time__gte=timezone.make_aware(datetime.combine(start_date, datetime.min.time())),
                    time__lte=timezone.make_aware(datetime.combine(end_date,   datetime.max.time())),
                )
                filename_suffix = f"{start_date}_to_{end_date}"
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        elif period == "today":
            queryset = queryset.filter(
                time__gte=timezone.make_aware(datetime.combine(today, datetime.min.time())),
                time__lte=timezone.make_aware(datetime.combine(today, datetime.max.time())),
            )
            filename_suffix = f"today_{today}"

        elif period == "last_week":
            seven_ago = today - timedelta(days=7)
            yesterday = today - timedelta(days=1)
            queryset  = queryset.filter(
                time__gte=timezone.make_aware(datetime.combine(seven_ago,  datetime.min.time())),
                time__lte=timezone.make_aware(datetime.combine(yesterday,  datetime.max.time())),
            )
            filename_suffix = f"last_7_days_{seven_ago}_to_{yesterday}"

        elif period == "current_month":
            month_start = today.replace(day=1)
            queryset    = queryset.filter(
                time__gte=timezone.make_aware(datetime.combine(month_start, datetime.min.time())),
                time__lte=timezone.make_aware(datetime.combine(today,       datetime.max.time())),
            )
            filename_suffix = today.strftime("%B_%Y")

        else:
            filename_suffix = "all_violations"

        violations = queryset.select_related("pipeline", "camera", "ml_model").order_by("-time")

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Violations"

        header_font      = Font(bold=True, color="FFFFFF", size=12)
        header_fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "ID", "Violation Type", "Date", "Time", "Plate Number",
            "Camera", "ML Model", "Pipeline", "Total Detections",
            "Detection Labels", "Created At",
        ]

        for col, header in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col)
            cell.value     = header
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment

        for row, v in enumerate(violations, 2):
            labels = [
                f"{d.get('label','unknown')} ({d.get('confidence',0):.2f})"
                for d in (v.detections or [])
            ]
            ws.cell(row=row, column=1,  value=v.id)
            ws.cell(row=row, column=2,  value=v.violation_type)
            ws.cell(row=row, column=3,  value=v.time.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4,  value=v.time.strftime("%H:%M:%S"))
            ws.cell(row=row, column=5,  value=v.plate_number or "N/A")
            ws.cell(row=row, column=6,  value=str(v.camera))
            ws.cell(row=row, column=7,  value=v.ml_model.name if v.ml_model else "N/A")
            ws.cell(row=row, column=8,  value=str(v.pipeline))
            ws.cell(row=row, column=9,  value=len(v.detections or []))
            ws.cell(row=row, column=10, value=", ".join(labels) if labels else "None")
            ws.cell(row=row, column=11, value=v.created_at.strftime("%Y-%m-%d %H:%M:%S"))

        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_len    = max(
                (len(str(c.value)) for c in ws[col_letter] if c.value),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="violations_{filename_suffix}.xlsx"'
        )
        wb.save(response)
        return response


# ── Period Report ──────────────────────────────────────────────────────────────
class ViolationPeriodReportView(APIView):
    """GET /violations/period-report/?period=today|last_week|current_month"""

    def _build_ranges(self, period, today):
        make = lambda d, t: timezone.make_aware(datetime.combine(d, t))

        if period == "today":
            return (
                make(today,                datetime.min.time()),
                make(today,                datetime.max.time()),
                make(today - timedelta(1), datetime.min.time()),
                make(today - timedelta(1), datetime.max.time()),
            )

        if period == "last_week":
            seven_ago = today - timedelta(days=7)
            yesterday = today - timedelta(days=1)
            fourteen  = today - timedelta(days=14)
            return (
                make(seven_ago,                datetime.min.time()),
                make(yesterday,                datetime.max.time()),
                make(fourteen,                 datetime.min.time()),
                make(seven_ago - timedelta(1), datetime.max.time()),
            )

        if period == "current_month":
            m_start = today.replace(day=1)
            if m_start.month == 1:
                prev_m = m_start.replace(year=m_start.year - 1, month=12, day=1)
            else:
                prev_m = m_start.replace(month=m_start.month - 1, day=1)
            prev_m_last = prev_m.replace(day=monthrange(prev_m.year, prev_m.month)[1])
            return (
                make(m_start,     datetime.min.time()),
                make(today,       datetime.max.time()),
                make(prev_m,      datetime.min.time()),
                make(prev_m_last, datetime.max.time()),
            )

        return None

    def get(self, request):
        period = request.query_params.get("period", "today")
        today  = timezone.now().date()

        ranges = self._build_ranges(period, today)
        if ranges is None:
            return Response(
                {"error": "Invalid period. Use: today | last_week | current_month"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cur_start, cur_end, prev_start, prev_end = ranges
        cur_qs    = Violation.objects.filter(time__gte=cur_start, time__lte=cur_end)
        cur_count = cur_qs.count()
        prev_count = Violation.objects.filter(
            time__gte=prev_start, time__lte=prev_end
        ).count()

        if prev_count == 0:
            trend_pct = 100.0 if cur_count > 0 else 0.0
            trend_dir = "up"  if cur_count > 0 else "neutral"
        else:
            raw       = ((cur_count - prev_count) / prev_count) * 100
            trend_pct = round(abs(raw), 1)
            trend_dir = "up" if raw > 0 else ("down" if raw < 0 else "neutral")

        breakdown_qs = (
            cur_qs.values("violation_type")
                  .annotate(count=Count("id"))
                  .order_by("-count")
        )
        breakdown = [
            {
                "violation_type": item["violation_type"],
                "count":          item["count"],
                "percentage":     round(item["count"] / cur_count * 100) if cur_count else 0,
            }
            for item in breakdown_qs
        ]

        time_fmt  = "%H:%M:%S" if period == "today" else ("%a %H:%M" if period == "last_week" else "%b %d")
        recent_qs = cur_qs.select_related("camera", "ml_model", "pipeline").order_by("-time")[:5]
        recent = [
            {
                "id":             v.id,
                "violation_type": v.violation_type,
                "plate_number":   v.plate_number or "N/A",
                "camera":         str(v.camera),
                "time":           v.time.strftime(time_fmt),
            }
            for v in recent_qs
        ]

        return Response({
            "period":            period,
            "total_count":       cur_count,
            "previous_count":    prev_count,
            "trend_percent":     trend_pct,
            "trend_direction":   trend_dir,
            "breakdown":         breakdown,
            "recent_violations": recent,
        })
